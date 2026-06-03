# %%------------   IMPORT CLASSES                         -----------------------------------------------------------> #
from main_code.sub_classes.single_phase import SPRotor, SPTeslaGeometry, SPTeslaOptions, SPStatorMil, SPStator
from main_code.base_classes import BaseTeslaTurbine
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from main_code.constants import CALCULATION_FOLDER
from tqdm import tqdm

# %%------------   IMPORT RESULTS                         -----------------------------------------------------------> #

CURRENT_FOLDER = os.path.join(CALCULATION_FOLDER, "04 - Experimental Check", "resources", "12_05_26")
RESULTS_FOLDER = os.path.join(CALCULATION_FOLDER, "04 - Experimental Check", "results")
file_path = os.path.join(CURRENT_FOLDER, "processed_labview_data.xlsx")

workbook = load_workbook(filename=file_path)

sheet = workbook['95bar']
n_rows = sheet.max_row - 1

headers = [cell.value for cell in sheet[1]]
data_dict = {header: np.empty(n_rows) for header in headers}
for header in headers:
    data_dict[header][:] = np.nan

n_row = 0
for row in sheet.iter_rows(min_row=2, values_only=True):

    for key, cell_value in zip(headers, row):

        data_dict[key][n_row] = cell_value

    n_row += 1

workbook.close()

thermodynamic_power = np.load(os.path.join(RESULTS_FOLDER, "power.npy"))

# %%------------   MAIN INPUT DATA                         ----------------------------------------------------------> #

P_inlet = []
P_outlet = []
T_inlet = []
rpm = []
T_outlet = []

for test in range(len(data_dict["P_Mandata"])):

    P_inlet.append(data_dict["P_Mandata"][test] * 100000)

    P_outlet.append(data_dict["P_Ritorno"][test] * 100000)

    T_inlet.append(data_dict["TC_Mandata"][test] + 273.15)

    rpm.append(data_dict["velocity measured"][test])

    T_outlet.append(data_dict["TC_Ritorno"][test] + 273.15)

n_packs = 33

# %%------------   LOSSES                         ----------------------------------------------------------> #
# The losses are derived from experimental data obtained rotating the turbine in the pressurized fluid without flow.

def rotor_exp_losses(rpm):

    c1 =  0.000000001250
    c2 =  0.000000400129
    c3 = 0.015765371647
    c4 = - 0.462712185101

    W_loss = c1 * rpm**3 + c2 * rpm**2 + c3 * rpm + c4

    return W_loss

# %%------------   CALCULATION                             ----------------------------------------------------------> #
total_power = list()
mass_flow = list()
rotor_inlet_pressure = list()
nozzle_pressure = list()
outlet_temperature = list()
rotor_inlet_temperature = list()
nozzle_temperature = list()
rotor_losses = list()
thermal_loss = list()
static_mdh = list()
ts_efficiency = list()

for i in (range(len(P_inlet))):

    P_in = P_inlet[i]
    P_out = P_outlet[i]
    T_in = T_inlet[i]

    # INITIALIZING TURBINE CONDITIONS
    curr_geometry = SPTeslaGeometry()
    curr_options = SPTeslaOptions()
    curr_options.rotor.integr_variable = 0.04  # [-]
    curr_options.stator.metastability_check = False
    curr_options.rotor.sp_check = True
    curr_options.rotor.n_rotor = 2000
    curr_options.rotor.heat_exchange = False
    curr_options.rotor.tank_temp = T_outlet[i]

    # Main design Parameters
    curr_geometry.rotor.b_channel = 0.00005
    curr_geometry.rotor.d_ratio = 3  # [m]
    curr_geometry.d_main = 0.15  # [m]

    tt = BaseTeslaTurbine("CarbonDioxide", curr_geometry, curr_options, stator=SPStatorMil, rotor=SPRotor)

    tt.rotor.rpm = rpm[i]
    losses = rotor_exp_losses(rpm[i])

    Z_stat = 1
    tt.geometry.stator.Z_stat = Z_stat

    correction_factor = 1.
    throat_width = 0.00114 * correction_factor
    tt.geometry.throat_width = throat_width

    tt.geometry.disc_thickness = 0.0001
    tt.geometry.rotor.n_discs = 1
    tt.geometry.H_s = 0.00015
    # tt.m_dot_tot = m_rif
    tt.n_packs = n_packs
    tt.geometry.n_channels = tt.n_packs

    alpha_in = 89  # [°]
    tt.geometry.alpha1 = alpha_in

    tt.points[0].set_variable("P", P_in)
    tt.points[0].set_variable("T", T_in)
    tt.stator.stator_eff = 0.9
    tt.rotor.gap_losses_control = True
    tt.include_extra_losses = False

    tt.P_in = P_in
    tt.P_out = P_out
    tt.T_in = T_in

    tt.iterate_pressure()
    tt.evaluate_performances()
    rotor_array = tt.rotor.get_rotor_array()

    total_power.append(tt.power * n_packs)
    mass_flow.append(tt.m_dot_tot)
    rotor_inlet_pressure.append(tt.static_points[2].get_variable("P") / 100000)
    nozzle_pressure.append(tt.static_points[1].get_variable("P") / 100000)
    outlet_temperature.append(tt.static_points[3].get_variable("T") - 273.15)
    nozzle_temperature.append(tt.static_points[1].get_variable("T") - 273.15)
    rotor_inlet_temperature.append(tt.static_points[2].get_variable("T") - 273.15)
    rotor_losses.append(losses)
    thermal_loss.append(tt.rotor.thermal_power)
    static_mdh.append(tt.stator.m_dot_s * tt.n_packs * (tt.static_points[0].get_variable("h") - tt.static_points[3].get_variable("h")))
    ts_efficiency.append(tt.Eta_tesla_ts)

    print("=" * 70)
    print(f"Test {i+1}/{len(P_inlet)} - Inlet: Pressure = {tt.static_points[0].get_variable('P') / 100000:.2f} bar,"
          f" Temperature = {tt.static_points[0].get_variable('T') - 273.15:.2f} °C,"
          f" Enthalpy = {tt.static_points[0].get_variable('h'):.2f} W")
    print(f"Test {i+1}/{len(P_inlet)} - Nozzle: Pressure = {tt.static_points[1].get_variable('P') / 100000:.2f} bar,"
          f" Temperature = {tt.static_points[1].get_variable('T') - 273.15:.2f} °C,"
          f" Enthalpy = {tt.static_points[1].get_variable('h'):.2f} W")
    print(f"Test {i+1}/{len(P_inlet)} - Rotor Inlet: Pressure = {tt.static_points[2].get_variable('P') / 100000:.2f} bar,"
          f" Temperature = {tt.static_points[2].get_variable('T') - 273.15:.2f} °C,"
          f" Enthalpy = {tt.static_points[2].get_variable('h'):.2f} W")
    print(f"Test {i+1}/{len(P_inlet)} - Outlet: Pressure = {tt.static_points[3].get_variable('P') / 100000:.2f} bar,"
          f" Temperature = {tt.static_points[3].get_variable('T') - 273.15:.2f} °C,"
          f" Enthalpy = {tt.static_points[3].get_variable('h'):.2f} W")
    print(f"Test {i+1}/{len(P_inlet)} -  Total Inlet: Pressure = {tt.points[0].get_variable('P') / 100000:.2f} bar,"
          f" Temperature = {tt.points[0].get_variable('T') - 273.15:.2f} °C,"
          f" Enthalpy = {tt.points[0].get_variable('h'):.2f} W")
    print(f"Test {i+1}/{len(P_inlet)} -  Total Nozzle: Pressure = {tt.points[1].get_variable('P') / 100000:.2f} bar,"
          f" Temperature = {tt.points[1].get_variable('T') - 273.15:.2f} °C,"
          f" Enthalpy = {tt.points[1].get_variable('h'):.2f} W")
    print(f"Test {i+1}/{len(P_inlet)} - Total Rotor Inlet: Pressure = {tt.points[2].get_variable('P') / 100000:.2f} bar,"
          f" Temperature = {tt.points[2].get_variable('T') - 273.15:.2f} °C,"
          f" Enthalpy = {tt.points[2].get_variable('h'):.2f} W")
    print(f"Test {i+1}/{len(P_inlet)} - Total Outlet: Pressure = {tt.points[3].get_variable('P') / 100000:.2f} bar,"
          f" Temperature = {tt.points[3].get_variable('T') - 273.15:.2f} °C,"
          f" Enthalpy = {tt.points[3].get_variable('h'):.2f} W")
    print("=" * 70)


# %%------------             COMPARISON PLOT                        -------------------------------------------------> #

fig, axs = plt.subplots(2,2, figsize=(9, 6), constrained_layout=True)

net_power = np.array(data_dict["power"][:]) + np.array(rotor_losses)

axs[1, 0].plot(rpm, total_power, c='darkblue', label="Calculated", ls='-', lw='0.5', marker='o', ms='3')
axs[1, 0].plot(rpm, data_dict["power"][:], c='darkred', label="Measured", ls='-', lw='0.5', marker='o', ms='3')
axs[1, 0].plot(rpm, thermodynamic_power, c='black', label=r"$m \cdot \Delta h$ Measured", ls='--', lw='1.5')
axs[1, 0].plot(rpm, static_mdh, c='darkblue', label=r"$m \cdot \Delta h$ Calculated", ls='--', lw='1.5')
axs[1, 0].set_xlabel("Rotational Speed [rpm]")
axs[1, 0].set_ylabel("Power [W]")
axs[1, 0].legend()
axs[1, 0].grid()

axs[0, 1].plot(rpm, mass_flow, c='darkblue', label="Calculated", ls='-', lw='0.5', marker='o', ms='3')
axs[0, 1].plot(rpm, data_dict["PORTATA"][:], c='darkred', label="Measured", ls='-', lw='0.5', marker='o', ms='3')
axs[0, 1].set_xlabel("Rotational Speed [rpm]")
axs[0, 1].set_ylabel("Mass Flow Rate [kg/s]")
axs[0, 1].legend()
axs[0, 1].grid()

# axs[0, 0].plot(rpm, rotor_inlet_pressure, c='darkblue', label="Calculated", ls='-', lw='0.5', marker='o', ms='3')
# axs[0, 0].plot(rpm, data_dict["P_Nozzle"][:], c='darkred', label="Measured", ls='-', lw='0.5', marker='o', ms='3')
# axs[0, 0].set_xlabel("Rotational Speed [rpm]")
# axs[0, 0].set_ylabel("Nozzle pressure [bar]")
# axs[0, 0].legend()
# axs[0, 0].grid()

axs[0, 0].plot(rpm, rotor_inlet_temperature, c='darkblue', label="Calculated", ls='-', lw='0.5', marker='o', ms='3')
axs[0, 0].plot(rpm, data_dict["T_Nozzle"][:], c='darkred', label="Measured", ls='-', lw='0.5', marker='o', ms='3')
axs[0, 0].set_xlabel("Rotational Speed [rpm]")
axs[0, 0].set_ylabel("Nozzle temperature [°C]")
axs[0, 0].legend()
axs[0, 0].grid()

axs[1, 1].plot(rpm, outlet_temperature, c='darkblue', label="Calculated", ls='-', lw='0.5', marker='o', ms='3')
axs[1, 1].plot(rpm, data_dict["TC_Ritorno"][:], c='darkred', label="Measured", ls='-', lw='0.5', marker='o', ms='3')
axs[1, 1].set_xlabel("Rotational Speed [rpm]")
axs[1, 1].set_ylabel("Outlet Temperature [°C]")
axs[1, 1].legend()
axs[1, 1].grid()

plt.savefig(os.path.join(RESULTS_FOLDER, "comparison_plot.png"), dpi=300)
plt.show()

# %%------------             AREA CORRECTION                        -------------------------------------------------> #
ratio = []

for i in range(len(mass_flow)):
    ratio.append(mass_flow[i] / data_dict["PORTATA"][i])

avg_ratio = np.mean(ratio)