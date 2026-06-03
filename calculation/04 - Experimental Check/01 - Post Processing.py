# %%------------   IMPORT CLASSES                         -----------------------------------------------------------> #
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from main_code.constants import CALCULATION_FOLDER
from tqdm import tqdm
from REFPROPConnector import ThermodynamicPoint as TP

# %%------------   IMPORT RESULTS                         -----------------------------------------------------------> #

CURRENT_FOLDER = os.path.join(CALCULATION_FOLDER, "04 - Experimental Check", "resources", "12_05_26")
file_path = os.path.join(CURRENT_FOLDER, "labviewdata.xlsx")

workbook = load_workbook(filename=file_path)

sheet = workbook['Seconda Misura']
n_rows = sheet.max_row - 1

headers = [cell.value for cell in sheet[1]]
data_dict = {header: np.empty(n_rows) for header in headers}
for header in headers:
    data_dict[header][:] = np.nan

n_row = 0
for row in sheet.iter_rows(min_row=2, values_only=True):

    for key, cell_value in zip(headers, row):

        data_dict[key][n_row] = cell_value

    n_row += 1

workbook.close()

# %%------------   Processing Results        -----------------------------------------------------------> #

P_inlet = data_dict["P_Mandata"] * 100000
T_inlet = data_dict["TC_Mandata"] + 273.15
P_outlet = data_dict["P_Ritorno"] * 100000
T_outlet = data_dict["TC_Ritorno"] + 273.15
P_nozzle = data_dict["P_Nozzle"] * 100000
T_nozzle = data_dict["T_Nozzle"] + 273.15
flow_rate = data_dict["PORTATA"]
Trigger = data_dict["TRIGGER"]

h_inlet, s_inlet = [], []
h_outlet, s_outlet = [], []
h_nozzle, s_nozzle = [], []

unique_triggers = np.unique(Trigger)

results = {trigger: {"h_inlet": [], "s_inlet": [], "h_outlet": [], "s_outlet": [],
                     "h_nozzle": [], "s_nozzle": [], "flow_rate": []} for trigger in unique_triggers}

# Calcolo delle proprietà termodinamiche
for i in tqdm(range(len(P_inlet))):

    inlet_point = TP(["CarbonDioxide"], [1], unit_system="MASS BASE SI")
    inlet_point.set_variable("P", P_inlet[i])
    inlet_point.set_variable("T", T_inlet[i])
    h_inlet = inlet_point.get_variable("h")
    s_inlet = inlet_point.get_variable("s")

    outlet_point = TP(["CarbonDioxide"], [1], unit_system="MASS BASE SI")
    outlet_point.set_variable("P", P_outlet[i])
    outlet_point.set_variable("T", T_outlet[i])
    h_outlet = outlet_point.get_variable("h")
    s_outlet = outlet_point.get_variable("s")

    nozzle_point = TP(["CarbonDioxide"], [1], unit_system="MASS BASE SI")
    nozzle_point.set_variable("P", P_nozzle[i])
    nozzle_point.set_variable("T", T_nozzle[i])
    h_nozzle = nozzle_point.get_variable("h")
    s_nozzle = nozzle_point.get_variable("s")

    # Salva i risultati in base al trigger
    results[Trigger[i]]["h_inlet"].append(h_inlet)
    results[Trigger[i]]["s_inlet"].append(s_inlet)
    results[Trigger[i]]["h_outlet"].append(h_outlet)
    results[Trigger[i]]["s_outlet"].append(s_outlet)
    results[Trigger[i]]["h_nozzle"].append(h_nozzle)
    results[Trigger[i]]["s_nozzle"].append(s_nozzle)
    results[Trigger[i]]["flow_rate"].append(flow_rate[i])

# Calcolo delle medie per ogni trigger

averages = {
    trigger: {
        "h_avg_inlet": np.mean(results[trigger]["h_inlet"]),
        "s_avg_inlet": np.mean(results[trigger]["s_inlet"]),
        "h_avg_outlet": np.mean(results[trigger]["h_outlet"]),
        "s_avg_outlet": np.mean(results[trigger]["s_outlet"]),
        "h_avg_nozzle": np.mean(results[trigger]["h_nozzle"]),
        "s_avg_nozzle": np.mean(results[trigger]["s_nozzle"]),
        "flow_rate_avg": np.mean(results[trigger]["flow_rate"])
    }
    for trigger in unique_triggers
}

power = []

for trigger in unique_triggers:

    power.append(averages[trigger]["flow_rate_avg"] * (averages[trigger]["h_avg_inlet"]
                                                       - averages[trigger]["h_avg_outlet"]))

    print(f"Trigger: {trigger}")
    print(f"Average Inlet Enthalpy: {averages[trigger]['h_avg_inlet']:.2f} J/kg")
    print(f"Average Inlet Entropy: {averages[trigger]['s_avg_inlet']:.4f} J/kg-K")
    print(f"Average Outlet Enthalpy: {averages[trigger]['h_avg_outlet']:.2f} J/kg")
    print(f"Average Outlet Entropy: {averages[trigger]['s_avg_outlet']:.4f} J/kg-K")
    print(f"Average Nozzle Enthalpy: {averages[trigger]['h_avg_nozzle']:.2f} J/kg")
    print(f"Average Nozzle Entropy: {averages[trigger]['s_avg_nozzle']:.4f} J/kg-K")
    print(f"Average Flow Rate: {averages[trigger]['flow_rate_avg']:.4f} kg/s\n")
    print(f"Calculated Power: {power[-1]:.2f} W\n")
RESULTS_FOLDER = os.path.join(CALCULATION_FOLDER, "04 - Experimental Check", "results")
np.save(os.path.join(RESULTS_FOLDER, "power.npy"), power)

# %%------------   PLOT RESULTS        -----------------------------------------------------------> #

plt.plot(unique_triggers, power, marker='o')
plt.show()