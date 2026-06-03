# %%------------   IMPORT CLASSES                         -----------------------------------------------------------> #
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from main_code.constants import CALCULATION_FOLDER
from tqdm import tqdm
from REFPROPConnector import ThermodynamicPoint as TP
from uncertainties import ufloat

# %%------------   IMPORT RESULTS                         -----------------------------------------------------------> #

CURRENT_FOLDER = os.path.join(CALCULATION_FOLDER, "04 - Experimental Check", "resources", "12_05_26")
file_path = os.path.join(CURRENT_FOLDER, "labviewdata.xlsx")

workbook = load_workbook(filename=file_path)

sheet = workbook['Seconda Misura']
n_rows = sheet.max_row - 1

headers = [cell.value for cell in sheet[1]]
data_dict = {header: np.empty(n_rows) for header in headers}
for header in headers:
    data_dict[header][:] = np.nan

n_row = 0
for row in sheet.iter_rows(min_row=2, values_only=True):

    for key, cell_value in zip(headers, row):

        data_dict[key][n_row] = cell_value

    n_row += 1

workbook.close()

# %%------------   Processing Results (Monte Carlo Approach) ---------------------> #

N_SAMPLES = 500  # Numero di campionamenti per ogni punto sperimentale
unique_triggers = np.unique(data_dict["TRIGGER"])

# Inizializziamo il dizionario per i risultati
results = {trigger: {"h_inlet": [], "h_outlet": [], "flow_rate": [],
                     "s_inlet": [], "s_outlet": [], "p_outlet": []} for trigger in unique_triggers}

# Definiamo le incertezze (standard deviation)

std_P_inlet = 0.48 * 100000
std_T_inlet = 0.3
std_P_outlet = 0.16 * 100000
std_T_outlet = 0.3

for i in tqdm(range(len(data_dict["P_Mandata"])), desc="Processing Points"):

    # 1. Recupero i valori nominali dal data_dict
    p_in_nom = data_dict["P_Mandata"][i] * 100000
    t_in_nom = data_dict["TC_Mandata"][i] + 273.15
    p_out_nom = data_dict["P_Ritorno"][i] * 100000
    t_out_nom = data_dict["TC_Ritorno"][i] + 273.15

    # 2. Genero i campioni rumorosi (distribuzione normale)
    p_in_samples = np.random.normal(p_in_nom, std_P_inlet, N_SAMPLES)
    t_in_samples = np.random.normal(t_in_nom, std_T_inlet, N_SAMPLES)
    p_out_samples = np.random.normal(p_out_nom, std_P_outlet, N_SAMPLES)
    t_out_samples = np.random.normal(t_out_nom, std_T_outlet, N_SAMPLES)

    h_in_list, s_in_list = [], []
    h_out_list, s_out_list = [], []

    # 3. Ciclo sui campioni (Propagazione dell'errore tramite REFPROP)

    inlet_point = TP(["CarbonDioxide"], [1], unit_system="MASS BASE SI")
    outlet_point = TP(["CarbonDioxide"], [1], unit_system="MASS BASE SI")


    for j in range(N_SAMPLES):
        # Inlet
        inlet_point.set_variable("P", p_in_samples[j])
        inlet_point.set_variable("T", t_in_samples[j])
        h_in_list.append(inlet_point.get_variable("h"))
        s_in_list.append(inlet_point.get_variable("s"))

        # Outlet
        outlet_point.set_variable("P", p_out_samples[j])
        outlet_point.set_variable("T", t_out_samples[j])
        h_out_list.append(outlet_point.get_variable("h"))
        s_out_list.append(outlet_point.get_variable("s"))

    # Salviamo la media e la deviazione standard (opzionale) per questo punto
    # Qui salviamo le liste complete per poi fare la media sul trigger
    curr_trigger = data_dict["TRIGGER"][i]
    results[curr_trigger]["h_inlet"].extend(h_in_list)
    results[curr_trigger]["h_outlet"].extend(h_out_list)
    results[curr_trigger]["s_inlet"].extend(s_in_list)
    results[curr_trigger]["s_outlet"].extend(s_out_list)
    results[curr_trigger]["flow_rate"].append(data_dict["PORTATA"][i])
    results[curr_trigger]["p_outlet"].append(p_out_nom)

# %%------------   Calcolo Potenza con Incertezza --------------------------------> #

for trigger in unique_triggers:
    h_in_array = np.array(results[trigger]["h_inlet"])
    h_out_array = np.array(results[trigger]["h_outlet"])
    s_in_array = np.array(results[trigger]["s_inlet"])
    s_out_array = np.array(results[trigger]["s_outlet"])
    flow_avg = np.mean(results[trigger]["flow_rate"])

    is_outlet_point = TP(["CarbonDioxide"], [1], unit_system="MASS BASE SI")
    is_outlet_point.set_variable("P", np.mean(results[trigger]["p_outlet"]))
    is_outlet_point.set_variable("s", np.mean(s_in_array))
    h_is_mean = is_outlet_point.get_variable("h")

    # Calcolo differenza di entalpia media e sua incertezza
    dh = h_in_array - h_out_array
    dh_mean = np.mean(dh)
    dh_std = np.std(dh)

    calc_power = flow_avg * dh_mean
    power_uncertainty = flow_avg * dh_std  # Assumendo flow_rate senza incertezza

    efficiency = dh_mean / (np.mean(h_in_array) - h_is_mean)

    print(f"Trigger: {trigger}")
    print(f"Potenza: {calc_power:.2f} ± {power_uncertainty:.2f} W")
    print(f"Efficienza: {efficiency:.4f}")
